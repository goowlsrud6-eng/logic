import hashlib
import re
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from django.db.models import Sum
from django.utils import timezone

from .models import DailyShipment, InboundSchedule, ProductMaster, ProductOptionMetric, UploadedFile

WEEK_RE = re.compile(r'(\d{4}-\d{4})')
SPECIAL_OPTION_RE = re.compile(r'★\s*\d+\s*차\s*')
OPTION_NORMALIZE_RE = re.compile(r'[\s/\\_\-\&\(\)\[\]\{\}\.,·]+')

COLUMN_ALIASES = {
    'order_number': ['발주번호', '발주 No', '발주NO', '오더번호', '일자-NO.', '일자NO', '일자-번호'],
    'product_code': ['상품코드', '이카운트코드'],
    'supplier_option_name': ['공급처옵션명', '공급처옵션', '품목코드'],
    'product_name': ['상품명', '품목명'],
    'option_name': ['옵션명', '옵션', '규격'],
    'available_stock': ['가용재고', '현재고'],
    'inbound_qty': ['총입고예정', '입고예정수량', '입고예정', '수량', '미구매수량'],
    'inbound_date': ['입고예정일', '입고일정', '일정', '출고일'],
    'memo': ['메모', '비고', '적요'],
    'delivery_qty': ['송장+배송', '배송'],
    'pending_qty': ['미출고', '접수'],
    'recent_week_sales': ['판매수량최근한주', '최근한주판매수량', '최근한주수량'],
    'total_sales': ['판매수량총판매', '총판매수량'],
    'sales_days': ['판매일수', '일수'],
    'open_date': ['오픈일', '판매시작일'],
    'status': ['상태'],
}


def normalize_header(value):
    text = str(value or '').replace('\n', '').replace('\r', '').strip().lower()
    return re.sub(r'[\s\-_/\.\(\)\[\]\{\}]+', '', text)


def normalize_lookup_text(value):
    return OPTION_NORMALIZE_RE.sub('', SPECIAL_OPTION_RE.sub('', str(value or '')).lower()).strip()


def clean_option_name(value):
    return SPECIAL_OPTION_RE.sub('', str(value or '')).strip()


def sha256_file(file_obj):
    digest = hashlib.sha256()
    for chunk in file_obj.chunks():
        digest.update(chunk)
    file_obj.seek(0)
    return digest.hexdigest()


def week_label_from_date(reference_date):
    if not reference_date:
        return ''
    monday = reference_date - timedelta(days=reference_date.weekday())
    friday = monday + timedelta(days=4)
    return f'{monday:%m%d}-{friday:%m%d}'


def infer_week_label(explicit_label, workbook_path, reference_date=None):
    if explicit_label:
        return explicit_label
    if reference_date:
        return week_label_from_date(reference_date)
    match = WEEK_RE.search(Path(workbook_path).name)
    return match.group(1) if match else ''


def find_header_row(raw_df):
    for idx in range(min(8, len(raw_df))):
        headers = [normalize_header(v) for v in raw_df.iloc[idx].tolist()]
        if any('상품명' in h or '품목명' in h for h in headers):
            return idx
    return 0


def build_column_map(columns):
    normalized = {col: normalize_header(col) for col in columns}
    mapped = {}
    for target, aliases in COLUMN_ALIASES.items():
        normalized_aliases = [normalize_header(alias) for alias in aliases]
        for alias in normalized_aliases:
            exact_match = next((col for col, header in normalized.items() if header == alias), None)
            if exact_match is not None:
                mapped[target] = exact_match
                break
        if target in mapped:
            continue
        for alias in normalized_aliases:
            partial_match = next((col for col, header in normalized.items() if alias in header), None)
            if partial_match is not None:
                mapped[target] = partial_match
                break
    return mapped


def as_number(value):
    if pd.isna(value):
        return 0.0
    numeric = pd.to_numeric(value, errors='coerce')
    return 0.0 if pd.isna(numeric) else float(numeric)


def parse_date(value, reference_date=None):
    """Parse common Korean Excel date inputs safely.

    Handles yyyymmdd, yyyy/mm/dd, yyyy-mm-dd, m/d, mm/dd and mmdd.
    Numeric yyyymmdd must be treated as a calendar date, not Unix nanoseconds.
    """
    if value is None or pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    base_year = (reference_date or timezone.localdate()).year
    text = str(value).strip()
    if not text:
        return None
    if text.endswith('.0'):
        text = text[:-2]

    compact = re.sub(r'[^0-9]', '', text)
    if len(compact) == 8:
        try:
            return datetime.strptime(compact, '%Y%m%d').date()
        except ValueError:
            pass
    if len(compact) == 2 and not re.match(r'^[12][0-9]{3}', compact):
        try:
            return date(base_year, int(compact[0]), int(compact[1]))
        except ValueError:
            pass
    if len(compact) in {3, 4} and not re.match(r'^[12][0-9]{3}', compact):
        compact = compact.zfill(4)
        try:
            return date(base_year, int(compact[:2]), int(compact[2:]))
        except ValueError:
            pass

    slash_match = re.search(r'(\d{1,2})[/-](\d{1,2})', text)
    if slash_match:
        try:
            return date(base_year, int(slash_match.group(1)), int(slash_match.group(2)))
        except ValueError:
            pass

    parsed = pd.to_datetime(text, errors='coerce')
    return parsed.date() if pd.notna(parsed) else None


def safe_weeks(stock, weekly_sales):
    if weekly_sales <= 0:
        return 0.0
    return round(stock / weekly_sales, 1)


def recent_sales_period_days(sales_days):
    if sales_days and sales_days > 0:
        return min(7, sales_days)
    return 7


def recent_weekly_rate(recent_sales, sales_days):
    period_days = recent_sales_period_days(sales_days)
    return (recent_sales / period_days * 7) if recent_sales > 0 and period_days > 0 else 0


def judge_stock_status(weeks):
    if weeks <= 0:
        return '판매없음'
    if weeks <= 4:
        return '재고부족위험'
    if weeks <= 8:
        return '주의'
    return '정상'


def normalize_sales_trend(value):
    if value == '판매 급상승':
        return '판매 급등'
    if value == '판매 급하락':
        return '판매 급감'
    return value or ''


def sales_trend_css_class(value):
    normalized = normalize_sales_trend(value)
    return {
        '판매 급등': 'trend-surge-strong',
        '판매 상승': 'trend-surge',
        '판매 하락': 'trend-drop',
        '판매 급감': 'trend-drop-strong',
    }.get(normalized, '')


def judge_sales_trend(current_weeks, previous_weeks):
    if current_weeks <= 0 or previous_weeks <= 0:
        return ''
    if current_weeks <= previous_weeks / 10:
        return '판매 급등'
    if current_weeks <= previous_weeks / 3:
        return '판매 상승'
    if current_weeks >= previous_weeks * 10:
        return '판매 급감'
    if current_weeks >= previous_weeks * 3:
        return '판매 하락'
    return ''


def metric_keys(product_code='', supplier_option_name='', product_name='', option_name=''):
    keys = []
    supplier = str(supplier_option_name or '').strip()
    code = str(product_code or '').strip()
    product = str(product_name or '').strip()
    normalized_option = normalize_lookup_text(option_name)
    if supplier:
        keys.append(('supplier', supplier))
    if code:
        keys.append(('code', code))
    if product and normalized_option:
        keys.append(('name', product, normalized_option))
    return keys or [('name', product, normalized_option)]


def metric_key(product_code='', supplier_option_name='', product_name='', option_name=''):
    return metric_keys(product_code, supplier_option_name, product_name, option_name)[0]


def first_lookup(lookup, product_code='', supplier_option_name='', product_name='', option_name='', default=0):
    for key in metric_keys(product_code, supplier_option_name, product_name, option_name):
        if key in lookup:
            return lookup[key]
    return default


def find_master_open_date(product_code, supplier_option_name, product_name, option_name):
    qs = ProductMaster.objects.filter(product_name=product_name)
    if supplier_option_name:
        found = qs.filter(supplier_option_name=supplier_option_name).first()
        if found and found.open_date:
            return found.open_date
    if product_code:
        found = qs.filter(product_code=product_code).first()
        if found and found.open_date:
            return found.open_date
    found = qs.filter(option_name=option_name).first() or qs.first()
    return found.open_date if found else None


def planned_inbound_by_key():
    today = timezone.localdate()
    rows = InboundSchedule.objects.filter(status=InboundSchedule.Status.PLANNED).values(
        'product_code', 'supplier_option_name', 'product_name', 'option_name'
    ).annotate(quantity=Sum('quantity'))
    result = defaultdict(float)
    for row in rows:
        for key in metric_keys(row['product_code'], row['supplier_option_name'], row['product_name'], row['option_name']):
            result[key] += row['quantity'] or 0
    return result


def previous_stock_file(current_file):
    qs = UploadedFile.objects.filter(
        status=UploadedFile.Status.COMPLETED,
        file_type__in=[UploadedFile.FileType.STOCK_SALES, UploadedFile.FileType.LEGACY],
        metrics__isnull=False,
    ).exclude(pk=current_file.pk).distinct()
    if current_file.reference_date:
        dated = qs.filter(reference_date__lt=current_file.reference_date).order_by('-reference_date', '-created_at').first()
        if dated:
            return dated
    return qs.filter(created_at__lt=current_file.created_at).order_by('-created_at').first() or qs.order_by('-reference_date', '-created_at').first()


def previous_recent_sales_by_key(current_file):
    previous_file = previous_stock_file(current_file)
    result = {}
    if not previous_file:
        return result
    for row in ProductOptionMetric.objects.filter(uploaded_file=previous_file):
        for key in metric_keys(row.product_code, row.supplier_option_name, row.product_name, row.option_name):
            result[key] = row.recent_week_sales
    return result


def autofill_supplier_options(items):
    lookup = {}
    for item in items:
        supplier = str(item.get('supplier_option_name') or '').strip()
        product = str(item.get('product_name') or '').strip()
        option = normalize_lookup_text(item.get('option_name'))
        if supplier and product and option:
            lookup[(product, option)] = supplier

    filled = 0
    failed = 0
    for item in items:
        if str(item.get('supplier_option_name') or '').strip():
            continue
        product = str(item.get('product_name') or '').strip()
        option = normalize_lookup_text(item.get('option_name'))
        supplier = lookup.get((product, option))
        if supplier:
            item['supplier_option_name'] = supplier
            filled += 1
        elif product and option:
            failed += 1
    return filled, failed


def build_metric(uploaded_file, item, week_label, source_sheet, inbound_qty, previous_recent_sales):
    product_code = str(item.get('product_code') or '').strip()
    supplier_option_name = str(item.get('supplier_option_name') or '').strip()
    product_name = str(item.get('product_name') or '').strip()
    option_name = clean_option_name(item.get('option_name'))
    stock = as_number(item.get('available_stock'))
    recent_sales = as_number(item.get('recent_week_sales'))
    total_sales = as_number(item.get('total_sales'))
    delivery_qty = as_number(item.get('delivery_qty'))
    pending_qty = as_number(item.get('pending_qty'))
    open_date = item.get('open_date') or find_master_open_date(product_code, supplier_option_name, product_name, option_name)
    reference_date = uploaded_file.reference_date or timezone.localdate()
    sales_days = max((reference_date - open_date).days, 1) if open_date else 0
    weekly_total_rate = (total_sales / sales_days * 7) if total_sales > 0 and sales_days > 0 else 0
    stock_after = stock + inbound_qty
    recent_rate = recent_weekly_rate(recent_sales, sales_days)
    current_recent = safe_weeks(stock, recent_rate)
    inbound_recent = safe_weeks(stock_after, recent_rate)
    current_total = safe_weeks(stock, weekly_total_rate)
    inbound_total = safe_weeks(stock_after, weekly_total_rate)
    prev_sales = first_lookup(previous_recent_sales, product_code, supplier_option_name, product_name, option_name, default=0)
    prev_weeks = safe_weeks(stock_after, prev_sales)
    return ProductOptionMetric(
        uploaded_file=uploaded_file,
        source_sheet=source_sheet,
        week_label=week_label,
        product_code=product_code,
        supplier_option_name=supplier_option_name,
        product_name=product_name,
        option_name=option_name,
        available_stock=stock,
        inbound_qty=inbound_qty,
        stock_after_inbound=stock_after,
        delivery_qty=delivery_qty,
        pending_qty=pending_qty,
        recent_week_sales=recent_sales,
        total_sales=total_sales,
        sales_days=sales_days,
        current_recent_weeks=current_recent,
        inbound_recent_weeks=inbound_recent,
        current_total_weeks=current_total,
        inbound_total_weeks=inbound_total,
        previous_inbound_recent_weeks=prev_weeks,
        status=judge_stock_status(inbound_recent or current_recent),
        sales_trend=judge_sales_trend(inbound_recent, prev_weeks),
    )


def finish_metric_upload(uploaded_file, items, source_sheet='기초파일'):
    filled, failed = autofill_supplier_options(items)
    inbound_lookup = planned_inbound_by_key()
    prev_lookup = previous_recent_sales_by_key(uploaded_file)
    week_label = infer_week_label(uploaded_file.week_label, uploaded_file.file.path, uploaded_file.reference_date)
    rows = []
    shipments = []

    for item in items:
        product_name = str(item.get('product_name') or '').strip()
        if not product_name or '합계' in product_name:
            continue
        inbound_qty = first_lookup(inbound_lookup, item.get('product_code'), item.get('supplier_option_name'), product_name, item.get('option_name'), default=0)
        metric = build_metric(uploaded_file, item, week_label, source_sheet, inbound_qty, prev_lookup)
        rows.append(metric)
        if uploaded_file.reference_date and metric.delivery_qty > 0:
            shipments.append(DailyShipment(
                uploaded_file=uploaded_file,
                delivery_date=uploaded_file.reference_date,
                product_code=metric.product_code,
                supplier_option_name=metric.supplier_option_name,
                product_name=metric.product_name,
                option_name=metric.option_name,
                quantity=metric.delivery_qty,
            ))

    ProductOptionMetric.objects.filter(uploaded_file=uploaded_file).delete()
    DailyShipment.objects.filter(uploaded_file=uploaded_file).delete()
    ProductOptionMetric.objects.bulk_create(rows)
    DailyShipment.objects.bulk_create(shipments)
    if uploaded_file.file_type != UploadedFile.FileType.LEGACY:
        uploaded_file.file_type = UploadedFile.FileType.STOCK_SALES
    uploaded_file.week_label = week_label
    uploaded_file.status = UploadedFile.Status.COMPLETED
    uploaded_file.message = f'{len(rows)}개 옵션 처리. 공급처옵션 자동 채움 성공 {filled}건 / 실패 {failed}건.'
    uploaded_file.save(update_fields=['file_type', 'week_label', 'status', 'message'])
    return len(rows)


def make_key(product_code, supplier_option_name, product_name, option_name):
    return (
        str(product_code or '').strip(),
        str(product_name or '').strip(),
        normalize_lookup_text(option_name),
    )


def parse_combined_single_sheet_workbook(uploaded_file):
    wb = load_workbook(uploaded_file.file.path, data_only=True)
    ws = wb.active
    items = {}
    for row_number in range(3, ws.max_row + 1):
        current = [ws.cell(row=row_number, column=1 + idx).value for idx in range(9)]
        recent = [ws.cell(row=row_number, column=11 + idx).value for idx in range(5)]
        total = [ws.cell(row=row_number, column=17 + idx).value for idx in range(5)]
        if not any(current + recent + total):
            continue
        if any(current):
            product_code, supplier, product_name, option_name, stock, pending, _invoice, delivery, invoice_delivery = current
            delivery_qty = invoice_delivery if invoice_delivery not in [None, ''] else delivery
            key = make_key(product_code, supplier, product_name, option_name)
            item = items.setdefault(key, {'product_code': product_code, 'supplier_option_name': supplier, 'product_name': product_name, 'option_name': option_name})
            if supplier and not item.get('supplier_option_name'):
                item['supplier_option_name'] = supplier
            item.update({'available_stock': stock, 'pending_qty': pending, 'delivery_qty': delivery_qty})
        if any(recent):
            product_code, supplier, product_name, option_name, qty = recent
            key = make_key(product_code, supplier, product_name, option_name)
            item = items.setdefault(key, {'product_code': product_code, 'supplier_option_name': supplier, 'product_name': product_name, 'option_name': option_name})
            if supplier and not item.get('supplier_option_name'):
                item['supplier_option_name'] = supplier
            item['recent_week_sales'] = qty
        if any(total):
            product_code, supplier, product_name, option_name, qty = total
            key = make_key(product_code, supplier, product_name, option_name)
            item = items.setdefault(key, {'product_code': product_code, 'supplier_option_name': supplier, 'product_name': product_name, 'option_name': option_name})
            if supplier and not item.get('supplier_option_name'):
                item['supplier_option_name'] = supplier
            item['total_sales'] = qty
    return finish_metric_upload(uploaded_file, list(items.values()), source_sheet=ws.title)


def parse_basic_inventory_workbook(uploaded_file):
    raw = pd.read_excel(uploaded_file.file.path, sheet_name=0, header=None)
    header_row = find_header_row(raw)
    df = pd.read_excel(uploaded_file.file.path, sheet_name=0, header=header_row).dropna(how='all')
    colmap = build_column_map(df.columns)
    required = ['product_name', 'available_stock', 'recent_week_sales', 'total_sales']
    missing = [name for name in required if name not in colmap]
    if missing:
        raise ValueError('필수 컬럼이 없습니다: ' + ', '.join(missing))
    items = []
    for _, row in df.iterrows():
        items.append({
            'product_code': row.get(colmap.get('product_code'), ''),
            'supplier_option_name': row.get(colmap.get('supplier_option_name'), ''),
            'product_name': row.get(colmap['product_name'], ''),
            'option_name': row.get(colmap.get('option_name'), ''),
            'available_stock': row.get(colmap['available_stock'], 0),
            'pending_qty': row.get(colmap.get('pending_qty'), 0),
            'delivery_qty': row.get(colmap.get('delivery_qty'), 0),
            'recent_week_sales': row.get(colmap['recent_week_sales'], 0),
            'total_sales': row.get(colmap['total_sales'], 0),
            'open_date': parse_date(row.get(colmap.get('open_date')), uploaded_file.reference_date) if 'open_date' in colmap else None,
        })
    return finish_metric_upload(uploaded_file, items, source_sheet='기초파일')


def parse_special_stock_workbook(uploaded_file):
    excel = pd.ExcelFile(uploaded_file.file.path)
    items = []
    for sheet_name in excel.sheet_names:
        if sheet_name == '옵션채우기' or '요약' in sheet_name or ('(' not in sheet_name and ')' not in sheet_name):
            continue
        raw = pd.read_excel(uploaded_file.file.path, sheet_name=sheet_name, header=None)
        if raw.empty:
            continue
        header_row = find_header_row(raw)
        df = pd.read_excel(uploaded_file.file.path, sheet_name=sheet_name, header=header_row).dropna(how='all')
        colmap = build_column_map(df.columns)
        if 'product_name' not in colmap or 'available_stock' not in colmap:
            continue
        for _, row in df.iterrows():
            items.append({
                'product_code': row.get(colmap.get('product_code'), ''),
                'supplier_option_name': row.get(colmap.get('supplier_option_name'), ''),
                'product_name': row.get(colmap['product_name'], ''),
                'option_name': row.get(colmap.get('option_name'), ''),
                'available_stock': row.get(colmap['available_stock'], 0),
                'pending_qty': row.get(colmap.get('pending_qty'), 0),
                'delivery_qty': row.get(colmap.get('delivery_qty'), 0),
                'recent_week_sales': row.get(colmap.get('recent_week_sales'), 0),
                'total_sales': row.get(colmap.get('total_sales'), 0),
                'open_date': parse_date(row.get(colmap.get('open_date')), uploaded_file.reference_date) if 'open_date' in colmap else None,
            })
    uploaded_file.file_type = UploadedFile.FileType.LEGACY
    uploaded_file.save(update_fields=['file_type'])
    return finish_metric_upload(uploaded_file, items, source_sheet='기존 특별재고')


def parse_product_master_workbook(uploaded_file):
    df = pd.read_excel(uploaded_file.file.path, sheet_name=0, header=0).dropna(how='all')
    colmap = build_column_map(df.columns)
    if 'product_name' not in colmap:
        raise ValueError('필수 컬럼이 없습니다: product_name')
    count = 0
    for _, row in df.iterrows():
        product_name = str(row.get(colmap['product_name'], '') or '').strip()
        if not product_name:
            continue
        ProductMaster.objects.update_or_create(
            product_code=str(row.get(colmap.get('product_code'), '') or '').strip() if 'product_code' in colmap else '',
            product_name=product_name,
            option_name=clean_option_name(row.get(colmap.get('option_name'), '')) if 'option_name' in colmap else '',
            defaults={
                'supplier_option_name': str(row.get(colmap.get('supplier_option_name'), '') or '').strip() if 'supplier_option_name' in colmap else '',
                'open_date': parse_date(row.get(colmap.get('open_date')), uploaded_file.reference_date) if 'open_date' in colmap else None,
            },
        )
        count += 1
    uploaded_file.file_type = UploadedFile.FileType.PRODUCT_MASTER
    uploaded_file.status = UploadedFile.Status.COMPLETED
    uploaded_file.message = f'{count}개 상품기본정보를 저장했습니다.'
    uploaded_file.save(update_fields=['file_type', 'status', 'message'])
    return count


def parse_inbound_schedule_workbook(uploaded_file):
    raw = pd.read_excel(uploaded_file.file.path, sheet_name=0, header=None)
    header_row = find_header_row(raw)
    df = pd.read_excel(uploaded_file.file.path, sheet_name=0, header=header_row).dropna(how='all')
    colmap = build_column_map(df.columns)
    required = ['product_name', 'inbound_qty']
    missing = [name for name in required if name not in colmap]
    if missing:
        raise ValueError('필수 컬럼이 없습니다: ' + ', '.join(missing))
    count = 0
    for _, row in df.iterrows():
        product_name = str(row.get(colmap['product_name'], '') or '').strip()
        qty = as_number(row.get(colmap['inbound_qty']))
        if not product_name or qty <= 0:
            continue
        inbound_date = parse_date(row.get(colmap.get('inbound_date')), uploaded_file.reference_date) if 'inbound_date' in colmap else None
        order_number = str(row.get(colmap.get('order_number'), '') or '').strip() if 'order_number' in colmap else ''
        product_code = str(row.get(colmap.get('product_code'), '') or '').strip() if 'product_code' in colmap else ''
        supplier = str(row.get(colmap.get('supplier_option_name'), '') or '').strip() if 'supplier_option_name' in colmap else ''
        option_name = clean_option_name(row.get(colmap.get('option_name'), '')) if 'option_name' in colmap else ''
        memo = str(row.get(colmap.get('memo'), '') or '').strip() if 'memo' in colmap else ''
        status_label = str(row.get(colmap.get('status'), '') or '').strip() if 'status' in colmap else ''
        status = {
            '완료': InboundSchedule.Status.COMPLETED,
            '취소': InboundSchedule.Status.CANCELED,
            '예정': InboundSchedule.Status.PLANNED,
        }.get(status_label, InboundSchedule.Status.PLANNED)
        base_qs = InboundSchedule.objects.filter(
            order_number=order_number,
            supplier_option_name=supplier,
            product_name=product_name,
            option_name=option_name,
        )
        if memo:
            target = base_qs.filter(memo=memo).first()
        else:
            same_option = list(base_qs[:2])
            target = same_option[0] if len(same_option) == 1 else base_qs.filter(inbound_date=inbound_date).first()

        defaults = {
            'uploaded_file': uploaded_file,
            'order_number': order_number,
            'product_code': product_code,
            'inbound_date': inbound_date,
            'quantity': qty,
            'memo': memo,
            'status': status,
            'is_completed': status == InboundSchedule.Status.COMPLETED,
        }
        if target:
            for field, value in defaults.items():
                setattr(target, field, value)
            target.save(update_fields=list(defaults.keys()) + ['updated_at'])
        else:
            InboundSchedule.objects.create(
                supplier_option_name=supplier,
                product_name=product_name,
                option_name=option_name,
                **defaults,
            )
        count += 1
    uploaded_file.file_type = UploadedFile.FileType.INBOUND_SCHEDULE
    uploaded_file.status = UploadedFile.Status.COMPLETED
    uploaded_file.message = f'{count}개 입고예정 일정을 저장/수정했습니다.'
    uploaded_file.save(update_fields=['file_type', 'status', 'message'])
    return count
